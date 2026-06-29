"""
File Handler
Manages file uploads, text extraction, and file operations.
"""
import os
from flask import current_app
from werkzeug.utils import secure_filename

from app.utils.helpers import generate_unique_filename


def allowed_file(filename, category='all'):
    """Check if filename has an allowed extension."""
    if '.' not in filename:
        return False
    ext = filename.rsplit('.', 1)[-1].lower()
    allowed = current_app.config.get('ALLOWED_EXTENSIONS', {}).get(category, set())
    return ext in allowed


def save_upload(file, subfolder='general'):
    """Save an uploaded file and return (saved_filename, file_path, file_size)."""
    if not file or not file.filename:
        return None, None, None

    original_name = secure_filename(file.filename)
    unique_name = generate_unique_filename(original_name)
    upload_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], subfolder)
    os.makedirs(upload_dir, exist_ok=True)

    file_path = os.path.join(upload_dir, unique_name)
    file.save(file_path)
    file_size = os.path.getsize(file_path)

    return unique_name, file_path, file_size


def delete_upload(file_path):
    """Safely delete an uploaded file."""
    if file_path and os.path.exists(file_path):
        try:
            os.remove(file_path)
            return True
        except OSError:
            return False
    return False


def get_file_extension(filename):
    """Extract file extension in lowercase."""
    if '.' in filename:
        return filename.rsplit('.', 1)[-1].lower()
    return ''


def extract_text_from_file(file_path):
    """Extract readable text from uploaded documents for AI context."""
    ext = get_file_extension(file_path)
    try:
        if ext == 'txt' or ext == 'md':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()

        elif ext == 'pdf':
            from PyPDF2 import PdfReader
            reader = PdfReader(file_path)
            text_parts = []
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
            return '\n'.join(text_parts)

        elif ext == 'docx':
            from docx import Document
            doc = Document(file_path)
            return '\n'.join(p.text for p in doc.paragraphs if p.text.strip())

        elif ext == 'csv':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()

        elif ext in ('xlsx', 'xls'):
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True)
            text_parts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = ' | '.join(str(cell) for cell in row if cell is not None)
                    if row_text.strip():
                        text_parts.append(row_text)
            return '\n'.join(text_parts)

    except Exception:
        return ''

    return ''
